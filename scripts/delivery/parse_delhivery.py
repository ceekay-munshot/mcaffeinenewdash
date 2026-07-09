# Parse the client-provided Delhivery Tracxn financials export into a clean
# data/clean/delivery.json — full 12-yr revenue + net-profit trend, an 11-yr
# DSO & EBITDA-margin trend, latest ratios, and the delivery-partner roster.
#   python3 scripts/delivery/parse_delhivery.py
import openpyxl, json, re

XLSX = "data/raw/Delivery Partners/Delhivery/Delhivery-Financials-FY2013-FY2024.xlsx"
wb = openpyxl.load_workbook(XLSX, data_only=True)
rows = lambda name: list(wb[name].iter_rows(values_only=True))

def fy_full(s):  # "24-25" -> "2024-25", "13-14" -> "2013-14"
    m = re.search(r"(\d{2})-(\d{2})", s)
    a, b = m.group(1), m.group(2)
    return f"20{a}-{b}"

def first_val(sheet, label, exact=False):
    lab = label.lower()
    for r in rows(sheet):
        if not (r and isinstance(r[0], str)):
            continue
        cell = r[0].strip().lower()
        if (cell == lab) if exact else cell.startswith(lab):
            v = r[-1]
            if isinstance(v, (int, float)):
                return float(v)
    return None

# 12-yr revenue + net profit, one per FY sheet
trend = []
for name in wb.sheetnames:
    m = re.match(r"\d+ FY (\d{2}-\d{2})", name)
    if not m:
        continue
    fy = fy_full(m.group(1))
    rev = first_val(name, "Total revenue")
    npf = first_val(name, "Total profit (loss) for period", exact=True)
    if npf is None:  # earliest years: approximate net result as revenue - expenses
        exp = first_val(name, "Total expenses")
        npf = (rev - exp) if (rev is not None and exp is not None) else None
    trend.append({"fy": fy, "revenueINR": rev, "netProfitINR": npf})
trend.sort(key=lambda t: t["fy"])

# 11-yr ratio trend (DSO, EBITDA margin) from the Ratios sheet
ratio_years, dso_row, ebitda_row = [], [], []
for r in rows("2 Ratios"):
    if not r or r[0] is None:
        continue
    k = str(r[0]).strip()
    if k == "Metrics":
        ratio_years = [fy_full(str(c)) for c in r[1:] if c]
    elif k == "Days Sales Outstanding":
        dso_row = list(r[1:])
    elif k == "EBITDA Margin":
        ebitda_row = list(r[1:])
ratioTrend = []
for i, fy in enumerate(ratio_years):
    dso = dso_row[i] if i < len(dso_row) else None
    eb = ebitda_row[i] if i < len(ebitda_row) else None
    ratioTrend.append({
        "fy": fy,
        "dso": round(dso, 1) if isinstance(dso, (int, float)) else None,
        "ebitdaMarginPct": round(eb * 100, 1) if isinstance(eb, (int, float)) else None,
    })
ratioTrend.sort(key=lambda t: t["fy"])

latest = trend[-1]
latestRatio = ratioTrend[-1]
delhivery = {
    "latestFY": latest["fy"],
    "revenueINR": latest["revenueINR"],
    "netProfitINR": latest["netProfitINR"],
    "dso": latestRatio["dso"],
    "ebitdaMarginPct": latestRatio["ebitdaMarginPct"],
    "trend": trend,
    "ratioTrend": ratioTrend,
}

ent = json.load(open("data/clean/entities.json"))
LISTED = {"Delhivery"}
partners = [
    {"brand": e["brand"], "legalName": e.get("legalName"), "cin": e.get("cin"),
     "coverage": e.get("coverage"), "listed": e["brand"] in LISTED}
    for e in ent["entities"] if e["category"] == "Delivery Partners"
]

json.dump({"partners": partners, "delhivery": delhivery},
          open("data/clean/delivery.json", "w"), indent=2)
print("wrote data/clean/delivery.json | trend years:", len(trend), "| ratio years:", len(ratioTrend))
print("revenue:", [(t["fy"], round(t["revenueINR"]/1e7) if t["revenueINR"] else None) for t in trend])
print("dso:", [(t["fy"], t["dso"]) for t in ratioTrend])
